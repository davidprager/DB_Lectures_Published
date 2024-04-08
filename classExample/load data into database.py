'''
This program loads data from Excel into database
'''

import mysql.connector
from openpyxl import load_workbook

# Create workbook object
book = load_workbook(filename = "Class.xlsx", data_only= True)

# Create worksheet object
data = book["Class"]

# Create connection object
mydb= mysql.connector.connect(
    host="localhost",
    user = "root",
    password = "dap123",
    database= "class"
)

# Create cursor
cursor = mydb.cursor()

# Define query
query = ("INSERT INTO class (name, age, gender, height, weight) VALUES (%s, %s, %s, %s, %s)")

# Process rows from Excel spreadsheet
data_row = 2
while data.cell(row=data_row, column = 1).value is not None:
    # Define variobles
    name = data.cell(row=data_row, column = 1).value
    age = data.cell(row=data_row, column = 2).value
    gender = data.cell(row=data_row, column = 3).value
    weight = data.cell(row=data_row, column = 4).value
    height = data.cell(row=data_row, column = 5).value
    record = (name, age, gender, weight, height)
    # Execute query
    cursor.execute(query, record)
    mydb.commit()

    data_row += 1
