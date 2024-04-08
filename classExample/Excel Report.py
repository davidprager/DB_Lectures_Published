'''
This program reads data from an Excel spreadsheet
'''

from openpyxl import load_workbook

# Create workbook object
book = load_workbook(filename = "Class.xlsx", data_only= True)

# Create worksheet object
data = book["Class"]

# Process all the data rows
male_count = 0
female_count = 0

data_row = 2
while data.cell(row=data_row, column = 1).value is not None:
    if data.cell(row=data_row, column = 3).value == "F":
        female_count += 1
    else:
        male_count += 1
    data_row += 1

print ("There are", male_count," male students and", female_count, "female students")
