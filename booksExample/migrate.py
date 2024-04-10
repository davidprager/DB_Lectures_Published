'''
This program loads book data from a spreadsheet into a database
'''

# import modules
import mysql.connector
from openpyxl import load_workbook

# Connect to books database
mydb = mysql.connector.connect(
  host="localhost",
  user="root",
  password="dap123",
    database = "books"
)
# Set up cursor
cursor = mydb.cursor()

# Open spreadsheet
book = load_workbook(filename = "books.xlsx", data_only=True)
data = book["Data"]

# Create query
query = "INSERT INTO book (title, author, isbn, category, price) values (%s, %s, %s ,%s, %s)"

# Iterate through spreadsheet
data_row = 2
while data.cell(row = data_row, column = 1).value is not None:
  title = data.cell(row = data_row, column = 1).value
  author = data.cell(row = data_row, column = 2).value
  ISBN = data.cell(row = data_row, column = 3).value
  category = data.cell(row = data_row, column = 4).value
  price = data.cell(row = data_row, column = 5).value

  # Set up tuple for insertiong
  record = (title, author, ISBN, category, price)

  # Insert data
  cursor.execute(query, record)
  mydb.commit()
  data_row += 1